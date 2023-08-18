import openpyxl
import pandas as pd
import os
import re

MainWorkBook_name = "流水信息"
DataWorkBook_name="流水识别数据源"

MainWorkBook = openpyxl.load_workbook(MainWorkBook_name+".xlsx")
DataWorkBook=openpyxl.load_workbook(DataWorkBook_name+".xlsx")
MainWorkSheet=MainWorkBook["sheet"]
DataWorkSheet_keyword=DataWorkBook["关键字"]
moneytransferlist=["归集","资金补足","上划","下拨","划转","手续费","服务费","账户维护费","结息","收息","利息","收费","出金","划拨","年费"]
cashiername="秦浩"
financername="王闯"
taxername="蒋允丽"
class TransactionRecord:
    def __init__(self,rownum,direction,abstract,remark,payname,paynum,receivenum,serialnum):
        self.rownum=rownum
        self.direction=direction
        self.abstract=abstract
        self.remark=remark
        self.payname=payname
        self.paynum=paynum
        self.receivenum=receivenum
        self.serialnum=serialnum
        self.responsibleperson=None
        self.matchmethod = None
        self.matchstat=False
    def renewresponsibleperson(self):
        MainWorkSheet.cell(self.rownum,21).value=self.responsibleperson
        MainWorkSheet.cell(self.rownum, 22).value = self.matchmethod
def Recongnition_serialnum(TransactionRecord):
    TransactionRecord.responsibleperson=lookup(DataWorkBook_name, "上一次数据", TransactionRecord.serialnum, "流水号", "匹配用户")
    if TransactionRecord.responsibleperson!=None:
        TransactionRecord.matchmethod = "Recongnition_serialnum"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True


def Recongnition_moneytransfer(TransactionRecord):
    if TransactionRecord.matchstat==True:
        return 0
    for keyword in moneytransferlist:
        if (keyword in TransactionRecord.abstract)or(keyword in TransactionRecord.remark):
            TransactionRecord.responsibleperson=cashiername
            break
    if TransactionRecord.responsibleperson!=None:
        TransactionRecord.matchmethod = "Recongnition_moneytransfer"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True

def Recongnition_peasantworker(TransactionRecord):
    if TransactionRecord.matchstat == True:
        return 0
    if TransactionRecord.direction=="收":
        TransactionRecord.responsibleperson = lookup(DataWorkBook_name, "农民工", TransactionRecord.receivenum, "银行账号", "匹配用户")
    else:
        TransactionRecord.responsibleperson = lookup(DataWorkBook_name, "农民工", TransactionRecord.paynum, "银行账号", "匹配用户")
    if TransactionRecord.responsibleperson!=None:
        TransactionRecord.matchmethod = "Recongnition_peasantworker"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True
    pass
def Recongnition_payer(TransactionRecord):
    if TransactionRecord.matchstat == True:
        return 0
    if TransactionRecord.direction=="收":
        TransactionRecord.responsibleperson = lookup(DataWorkBook_name, "付款方", TransactionRecord.payname, "付款户名", "匹配用户")
    if TransactionRecord.responsibleperson!=None:
        TransactionRecord.matchmethod = "Recongnition_payer"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True
def Recongnition_electricity(TransactionRecord):
    if TransactionRecord.matchstat == True:
        return 0
    if "电费" in TransactionRecord.abstract or "电力缴费" in TransactionRecord.abstract:
        elenum=int(numextract(TransactionRecord.abstract))
        TransactionRecord.responsibleperson=lookup(DataWorkBook_name, "电费", elenum, "号码", "匹配用户")
        if TransactionRecord.responsibleperson==None:
            TransactionRecord.responsibleperson="电费"
        TransactionRecord.matchmethod = "Recongnition_electricity"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True
def Recongnition_water(TransactionRecord):
    if TransactionRecord.matchstat == True:
        return 0
    if "水费" in TransactionRecord.abstract:
        TransactionRecord.responsibleperson="水费"
        TransactionRecord.matchmethod = "Recongnition_water"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True
def Recongnition_workinsurance(TransactionRecord):
    if TransactionRecord.matchstat == True:
        return 0
    if TransactionRecord.abstract=="工伤待遇":
        TransactionRecord.responsibleperson="社保"
        TransactionRecord.matchmethod = "Recongnition_workinsurance"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True
    pass
def Recongnition_tax(TransactionRecord):
    if TransactionRecord.matchstat == True:
        return 0
    if "企业所得税" in TransactionRecord.abstract:
        TransactionRecord.responsibleperson="刘文"
        TransactionRecord.matchmethod = "Recongnition_tax"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True
    elif "缴税" in TransactionRecord.abstract or "银联消费" in TransactionRecord.abstract:
        TransactionRecord.responsibleperson="税费"
        TransactionRecord.matchmethod = "Recongnition_tax"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True

def Recongnition_letterofguarantee(TransactionRecord):
    if TransactionRecord.matchstat == True:
        return 0
    if "保函" in TransactionRecord.remark:
        TransactionRecord.responsibleperson=financername
        TransactionRecord.matchmethod = "Recongnition_letterofguarantee"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True
def Recongnition_keyword(TransactionRecord):
    if TransactionRecord.matchstat == True:
        return 0
    for rownum in range(2,DataWorkSheet_keyword.max_row+1):
        if (DataWorkSheet_keyword.cell(rownum,1).value in TransactionRecord.abstract) or (DataWorkSheet_keyword.cell(rownum,1).value in TransactionRecord.remark):
            TransactionRecord.responsibleperson=DataWorkSheet_keyword.cell(rownum,2).value
            break
    if TransactionRecord.responsibleperson!=None:
        TransactionRecord.matchmethod = "Recongnition_keyword"
        TransactionRecord.renewresponsibleperson()
        TransactionRecord.matchstat=True
def sumRecongnition(TransactionRecord):
    Recongnition_serialnum(TransactionRecord)
    Recongnition_moneytransfer(TransactionRecord)
    Recongnition_peasantworker(TransactionRecord)
    Recongnition_payer(TransactionRecord)
    Recongnition_electricity(TransactionRecord)
    Recongnition_water(TransactionRecord)
    Recongnition_workinsurance(TransactionRecord)
    Recongnition_tax(TransactionRecord)
    Recongnition_letterofguarantee(TransactionRecord)
    Recongnition_keyword(TransactionRecord)
    if TransactionRecord.matchstat==False:
        TransactionRecord.responsibleperson = "？"
        TransactionRecord.matchmethod = "Recongnition_NoFound"
        TransactionRecord.renewresponsibleperson()
def TransactionRecord_Generate(rownum):#rownum,direction,abstract,remark,payname,paynum,receivenum,serialnum
    Record=TransactionRecord(rownum,
                             MainWorkSheet.cell(rownum,4).value,
                             MainWorkSheet.cell(rownum,6).value,
                             MainWorkSheet.cell(rownum,7).value,
                             MainWorkSheet.cell(rownum, 9).value,
                             MainWorkSheet.cell(rownum,10).value,
                             MainWorkSheet.cell(rownum,13).value,
                             str(MainWorkSheet.cell(rownum,24).value),
                             )
    return Record
def lookup(wbname,wsname,lookupvalue,lookupname,returnname):
    df = pd.read_excel(wbname+'.xlsx', sheet_name=wsname)
    try:
        result = df.loc[df[lookupname] == lookupvalue,returnname].values[0]
    except:
        result=None
    return result
def numextract(str):
    result=re.findall(r'\d+', str)[0]
    return result
def main():
    for rownum in range(2,MainWorkSheet.max_row+1):
        Record=TransactionRecord_Generate(rownum)
        sumRecongnition(Record)
        print("现在完成%s个，总共有%s个"%(rownum-1,MainWorkSheet.max_row-1))
    MainWorkBook.save(MainWorkBook_name+"完成.xlsx")

main()
